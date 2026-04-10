from __future__ import annotations

import argparse
import html
import json
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

import openpyxl
from playwright.sync_api import Page, Response, sync_playwright


HOME_MODULE = "首页"
SHEET_NAME = "全量测试用例"
SELECTED_CASE_IDS = ("TC-HOME-007", "TC-HOME-011", "TC-HOME-014")
DEFAULT_BASE_URL = "https://192.168.88.238:9608/#/login"
DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD = "admin@123"
DEFAULT_EXCEL_PATH = Path("testcases/zhdz-testcase.xlsx")
DEFAULT_ROUTE_TIMEOUT = 180
DEFAULT_CONTENT_TIMEOUT = 120
DEFAULT_MAX_RETRIES = 1
POLL_INTERVAL_SECONDS = 5


@dataclass
class CaseDefinition:
    case_id: str
    module: str
    feature: str
    description: str
    expected: str


@dataclass
class CaseResult:
    case_id: str
    case_name: str
    status: str
    duration_seconds: float
    expected: str
    actual: str
    page_url: str
    warnings: list[str] = field(default_factory=list)
    artifacts: dict[str, str] = field(default_factory=dict)


@dataclass
class WaitOutcome:
    success: bool
    waited_seconds: float
    details: str
    page_url: str
    text: str = ""
    polls: list[dict[str, object]] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="首页自动化独立脚本")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--username", default=DEFAULT_USERNAME)
    parser.add_argument("--password", default=DEFAULT_PASSWORD)
    parser.add_argument("--excel-path", default=str(DEFAULT_EXCEL_PATH))
    parser.add_argument("--report-dir", default="")
    parser.add_argument("--headless", default="true")
    parser.add_argument("--route-timeout", type=int, default=DEFAULT_ROUTE_TIMEOUT)
    parser.add_argument("--content-timeout", type=int, default=DEFAULT_CONTENT_TIMEOUT)
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES)
    return parser.parse_args()


def parse_bool(value: str | bool) -> bool:
    if isinstance(value, bool):
        return value
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(f"无法解析布尔值: {value}")


def build_report_dir(raw_report_dir: str) -> Path:
    if raw_report_dir:
        return Path(raw_report_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("reports") / "homepage_standalone" / timestamp


def write_text_file(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def log_step(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def load_home_cases(excel_path: Path) -> list[CaseDefinition]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"未找到sheet: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空，无法读取测试用例")

    current_module = ""
    selected_cases: list[CaseDefinition] = []
    for row in rows[1:]:
        # Excel 中“功能模块”列是合并展示风格，这里按常见手工用例表的方式向下继承模块名。
        module_value = row[2] if len(row) > 2 else None
        if module_value:
            current_module = str(module_value).strip()

        case_id = row[0] if len(row) > 0 else None
        if current_module != HOME_MODULE or case_id not in SELECTED_CASE_IDS:
            continue

        selected_cases.append(
            CaseDefinition(
                case_id=str(case_id),
                module=current_module,
                feature=str(row[6] or "").strip(),
                description=str(row[7] or "").strip(),
                expected=str(row[10] or "").strip(),
            )
        )

    missing = set(SELECTED_CASE_IDS) - {case.case_id for case in selected_cases}
    if missing:
        raise ValueError(f"Excel 中缺少目标用例: {sorted(missing)}")

    selected_cases.sort(key=lambda item: item.case_id)
    return selected_cases


def append_console_log(console_log_path: Path, level: str, text: str) -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with console_log_path.open("a", encoding="utf-8") as file:
        file.write(f"[{timestamp}] [{level}] {text}\n")


def body_text(page: Page) -> str:
    try:
        return page.evaluate("() => document.body ? document.body.innerText : ''")
    except Exception as exc:  # noqa: BLE001
        return f"[body.innerText failed] {exc}"


def current_hash(page: Page) -> str:
    try:
        return page.evaluate("() => location.hash")
    except Exception:  # noqa: BLE001
        return ""


def sanitize_filename(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", name)


def save_text_snapshot(page: Page, snapshot_dir: Path, case_id: str, attempt: int, stage: str) -> str:
    snapshot_path = snapshot_dir / f"{sanitize_filename(case_id)}_attempt{attempt}_{stage}.txt"
    write_text_file(snapshot_path, body_text(page))
    return str(snapshot_path)


def capture_screenshot(page: Page, screenshot_path: Path, label: str) -> str:
    # 慢环境下 full_page 截图可能因为字体/资源迟迟未就绪而超时，
    # 这里改成降级链路：先整页，再视口，最后写错误文件，但不抛异常中断测试主流程。
    screenshot_path.parent.mkdir(parents=True, exist_ok=True)
    screenshot_style = """
    * {
      animation: none !important;
      transition: none !important;
      caret-color: transparent !important;
    }
    video {
      visibility: hidden !important;
    }
    """
    full_page_error_text = ""
    try:
        page.screenshot(
            path=str(screenshot_path),
            full_page=True,
            timeout=30000,
            animations="disabled",
            caret="hide",
            scale="css",
            style=screenshot_style,
        )
        return str(screenshot_path)
    except Exception as full_page_exc:  # noqa: BLE001
        full_page_error_text = str(full_page_exc)
        log_step(f"{label} 整页截图失败，改为视口截图: {full_page_error_text}")

    try:
        page.screenshot(
            path=str(screenshot_path),
            full_page=False,
            timeout=15000,
            animations="disabled",
            caret="hide",
            scale="css",
            style=screenshot_style,
        )
        return str(screenshot_path)
    except Exception as viewport_exc:  # noqa: BLE001
        error_path = screenshot_path.with_name(f"{screenshot_path.stem}_error.txt")
        write_text_file(
            error_path,
            (
                f"{label} 截图失败\n"
                f"页面地址: {page.url}\n"
                f"整页截图错误: {full_page_error_text}\n"
                f"视口截图错误: {viewport_exc}\n"
            ),
        )
        log_step(f"{label} 视口截图也失败，已写入错误文件: {error_path}")
        return str(error_path)


def save_screenshot(page: Page, screenshot_dir: Path, case_id: str, attempt: int, stage: str) -> str:
    screenshot_path = screenshot_dir / f"{sanitize_filename(case_id)}_attempt{attempt}_{stage}.png"
    return capture_screenshot(page, screenshot_path, f"{case_id} attempt {attempt} {stage}")


def save_failure_logs(
    report_dir: Path,
    case_id: str,
    attempt: int,
    console_events: list[dict[str, object]],
    network_errors: list[dict[str, object]],
) -> dict[str, str]:
    artifacts: dict[str, str] = {}
    console_path = report_dir / f"{sanitize_filename(case_id)}_attempt{attempt}_console_errors.json"
    console_payload = [event for event in console_events if str(event.get("level")) in {"error", "pageerror"}]
    write_text_file(console_path, json.dumps(console_payload[-20:], ensure_ascii=False, indent=2))
    artifacts[f"attempt_{attempt}_console_errors"] = str(console_path)

    network_path = report_dir / f"{sanitize_filename(case_id)}_attempt{attempt}_network_errors.json"
    write_text_file(network_path, json.dumps(network_errors[-20:], ensure_ascii=False, indent=2))
    artifacts[f"attempt_{attempt}_network_errors"] = str(network_path)
    return artifacts


def wait_for_route_ready(page: Page, timeout_seconds: int) -> WaitOutcome:
    start = time.monotonic()
    polls: list[dict[str, object]] = []
    log_step(f"开始等待首页路由就绪，超时 {timeout_seconds}s。")
    while True:
        elapsed = round(time.monotonic() - start, 2)
        page_hash = current_hash(page)
        text = body_text(page)
        polls.append(
            {
                "elapsed_seconds": elapsed,
                "url": page.url,
                "hash": page_hash,
                "text_length": len(text),
            }
        )
        if page_hash == "#/globalOverview/index":
            log_step(f"首页路由已就绪，耗时 {elapsed}s。")
            return WaitOutcome(
                success=True,
                waited_seconds=elapsed,
                details="首页路由已就绪",
                page_url=page.url,
                text=text,
                polls=polls,
            )
        if elapsed >= timeout_seconds:
            log_step(f"首页路由等待失败，已等待 {elapsed}s，当前 hash={page_hash or '[empty]'}。")
            return WaitOutcome(
                success=False,
                waited_seconds=elapsed,
                details=f"{timeout_seconds}s 内未进入首页路由，当前 hash={page_hash or '[empty]'}",
                page_url=page.url,
                text=text,
                polls=polls,
            )
        log_step(
            f"等待首页路由中: {elapsed}/{timeout_seconds}s, 当前 hash={page_hash or '[empty]'}, 文本长度={len(text)}。"
        )
        # 映射环境很慢，固定轮询比依赖短超时断言更稳。
        page.wait_for_timeout(POLL_INTERVAL_SECONDS * 1000)


def wait_for_home_text_ready(page: Page, timeout_seconds: int) -> WaitOutcome:
    start = time.monotonic()
    polls: list[dict[str, object]] = []
    required_markers = ["首页", "风电机组信息", "光伏机组信息", "二氧化硫"]
    log_step(f"开始等待首页核心内容渲染，超时 {timeout_seconds}s。")
    while True:
        elapsed = round(time.monotonic() - start, 2)
        text = body_text(page)
        markers_ready = all(marker in text for marker in required_markers)
        polls.append(
            {
                "elapsed_seconds": elapsed,
                "url": page.url,
                "hash": current_hash(page),
                "text_length": len(text),
                "markers_ready": markers_ready,
            }
        )
        if markers_ready:
            log_step(f"首页核心内容已渲染，耗时 {elapsed}s。")
            return WaitOutcome(
                success=True,
                waited_seconds=elapsed,
                details="首页核心文本已渲染",
                page_url=page.url,
                text=text,
                polls=polls,
            )
        if elapsed >= timeout_seconds:
            log_step(f"首页核心内容等待失败，已等待 {elapsed}s，文本长度={len(text)}。")
            return WaitOutcome(
                success=False,
                waited_seconds=elapsed,
                details=f"{timeout_seconds}s 内首页核心文本未补齐",
                page_url=page.url,
                text=text,
                polls=polls,
            )
        log_step(
            f"等待首页核心内容中: {elapsed}/{timeout_seconds}s, 文本长度={len(text)}, 标记齐全={markers_ready}。"
        )
        # 首页先切路由、后补数据区，所以这里单独做第二阶段等待。
        page.wait_for_timeout(POLL_INTERVAL_SECONDS * 1000)


def wait_for_login_controls(page: Page, timeout_seconds: int = 60) -> None:
    start = time.monotonic()
    log_step(f"开始等待登录控件出现，超时 {timeout_seconds}s。")
    while True:
        inputs = page.locator("input").count()
        buttons = page.locator("button").count()
        if inputs >= 2 and buttons >= 1:
            log_step(f"登录控件已出现，检测到 {inputs} 个输入框、{buttons} 个按钮。")
            return
        if time.monotonic() - start >= timeout_seconds:
            raise RuntimeError("登录页未检测到用户名/密码输入框或登录按钮")
        log_step(f"等待登录控件中: 输入框={inputs}, 按钮={buttons}。")
        page.wait_for_timeout(POLL_INTERVAL_SECONDS * 1000)


def prepare_homepage(
    page: Page,
    base_url: str,
    username: str,
    password: str,
    route_timeout: int,
    content_timeout: int,
) -> tuple[WaitOutcome, WaitOutcome]:
    log_step(f"访问目标地址: {base_url}")
    page.goto(base_url, wait_until="domcontentloaded", timeout=180000)
    page.wait_for_timeout(15000)

    if current_hash(page) == "#/login":
        # 登录控件在慢环境下可能延后出现，先等控件再填值，避免刚到页就误判失败。
        log_step("检测到登录页，开始执行登录。")
        wait_for_login_controls(page)
        inputs = page.locator("input")
        buttons = page.locator("button")
        inputs.nth(0).fill(username, timeout=60000)
        inputs.nth(1).fill(password, timeout=60000)
        buttons.nth(0).click(timeout=60000)
        log_step("登录信息已提交。")
    else:
        log_step("未进入登录页，直接检查首页路由。")

    # 首页准备分成两个明确阶段，便于报告里定位是卡在登录/路由，还是卡在首页渲染。
    route_wait = wait_for_route_ready(page, route_timeout)
    if not route_wait.success:
        raise RuntimeError(f"登录/路由切换失败: {route_wait.details}")

    content_wait = wait_for_home_text_ready(page, content_timeout)
    if not content_wait.success:
        raise RuntimeError(f"首页核心内容未完成渲染: {content_wait.details}")

    return route_wait, content_wait


def extract_numeric_after(text: str, label: str) -> str | None:
    pattern = re.compile(re.escape(label) + r"[\s:：]*([0-9]+(?:\.[0-9]+)?)")
    match = pattern.search(text)
    if match:
        return match.group(1)
    return None


def require_numeric_after(text: str, label: str, scope_name: str) -> str:
    value = extract_numeric_after(text, label)
    if value is None:
        raise AssertionError(f"{scope_name}中“{label}”附近未提取到有效数值")
    return value


def extract_section(text: str, start_marker: str, end_markers: list[str]) -> str:
    start_index = text.find(start_marker)
    if start_index < 0:
        return ""
    end_index = len(text)
    for marker in end_markers:
        marker_index = text.find(marker, start_index + len(start_marker))
        if marker_index >= 0:
            end_index = min(end_index, marker_index)
    return text[start_index:end_index]


def summarize_runtime_warnings(
    console_events: list[dict[str, object]],
    network_errors: list[dict[str, object]],
) -> list[str]:
    warnings: list[str] = []
    if network_errors:
        recent_network = network_errors[-3:]
        warnings.append(
            "检测到 HTTP >=400 响应: "
            + "; ".join(f"{item['status']} {item['url']}" for item in recent_network)
        )
    error_events = [event for event in console_events if str(event.get("level")) in {"error", "pageerror"}]
    if error_events:
        recent_console = error_events[-3:]
        warnings.append(
            "检测到控制台错误: "
            + "; ".join(str(item["text"]).replace("\n", " ")[:180] for item in recent_console)
        )
    return warnings


def assert_home_generation(text: str) -> str:
    for label in ["日发电量", "月发电量", "年发电量", "实时功率"]:
        if label not in text:
            raise AssertionError(f"发电量总览缺少字段: {label}")
    if "万kWh" not in text or "MW" not in text:
        raise AssertionError("发电量总览缺少单位“万kWh”或“MW”")

    # 这里要求字段附近能抽取到数值，避免只渲染标签和单位时也被误判为通过。
    day_value = require_numeric_after(text, "日发电量", "发电量总览")
    month_value = require_numeric_after(text, "月发电量", "发电量总览")
    year_value = require_numeric_after(text, "年发电量", "发电量总览")
    power_value = require_numeric_after(text, "实时功率", "发电量总览")

    return (
        f"发电量总览已渲染，日发电量={day_value}万kWh，"
        f"月发电量={month_value}万kWh，年发电量={year_value}万kWh，"
        f"实时功率={power_value}MW。"
    )


def assert_turbine_and_pv_sections(text: str) -> str:
    if "风电机组信息" not in text or "光伏机组信息" not in text:
        raise AssertionError("风电机组信息区或光伏机组信息区未出现")

    wind_section = extract_section(text, "风电机组信息", ["光伏机组信息"])
    pv_section = extract_section(text, "光伏机组信息", ["集控整体监控", "二氧化硫"])
    if not wind_section or not pv_section:
        raise AssertionError("风电机组信息区或光伏机组信息区未完整渲染")

    wind_values = {
        "装机容量": require_numeric_after(wind_section, "装机容量", "风电机组信息区"),
        "平均风速": require_numeric_after(wind_section, "平均风速", "风电机组信息区"),
        "实时功率": require_numeric_after(wind_section, "实时功率", "风电机组信息区"),
        "装机台数": require_numeric_after(wind_section, "装机台数", "风电机组信息区"),
    }
    pv_values = {
        "装机容量": require_numeric_after(pv_section, "装机容量", "光伏机组信息区"),
        "辐照度": require_numeric_after(pv_section, "辐照度", "光伏机组信息区"),
        "实时功率": require_numeric_after(pv_section, "实时功率", "光伏机组信息区"),
        "装机台数": require_numeric_after(pv_section, "装机台数", "光伏机组信息区"),
    }

    for unit in ["MW", "m/s", "W/m²", "台"]:
        if unit not in text:
            raise AssertionError(f"首页缺少单位: {unit}")

    return (
        "风电与光伏区域同时可见，"
        f"风电装机容量={wind_values['装机容量']}MW，平均风速={wind_values['平均风速']}m/s，"
        f"光伏装机容量={pv_values['装机容量']}MW，辐照度={pv_values['辐照度']}W/m²。"
    )


def assert_energy_reduction(
    text: str,
    console_events: list[dict[str, object]],
    network_errors: list[dict[str, object]],
) -> tuple[str, list[str]]:
    for label in ["二氧化硫", "二氧化碳", "标准煤"]:
        if label not in text:
            raise AssertionError(f"节能减排区域缺少字段: {label}")
    if text.count("万吨/年") < 3:
        raise AssertionError("节能减排区域未检测到 3 个“万吨/年”单位")

    sulfur = require_numeric_after(text, "二氧化硫", "节能减排区域")
    carbon = require_numeric_after(text, "二氧化碳", "节能减排区域")
    coal = require_numeric_after(text, "标准煤", "节能减排区域")

    warnings = summarize_runtime_warnings(console_events, network_errors)
    actual = (
        f"节能减排区域已渲染，二氧化硫={sulfur}万吨/年，"
        f"二氧化碳={carbon}万吨/年，标准煤={coal}万吨/年。"
    )
    return actual, warnings


def run_case_with_retry(
    page: Page,
    case: CaseDefinition,
    report_dir: Path,
    console_events: list[dict[str, object]],
    network_errors: list[dict[str, object]],
    assertion: Callable[[str], str | tuple[str, list[str]]],
    base_url: str,
    username: str,
    password: str,
    route_timeout: int,
    content_timeout: int,
    max_retries: int,
) -> CaseResult:
    snapshots_dir = report_dir / "snapshots"
    screenshots_dir = report_dir / "screenshots"
    warnings: list[str] = []
    artifacts: dict[str, str] = {}
    start_time = time.monotonic()
    last_error = "未执行"

    for attempt in range(1, max_retries + 2):
        try:
            log_step(f"[{case.case_id}] 开始第 {attempt} 次执行: {case.feature}")
            # 每次执行前都确认首页还处于可断言状态，避免上一次失败把页面留在异常中间态。
            readiness = wait_for_home_text_ready(page, min(content_timeout, 60))
            if not readiness.success:
                raise AssertionError(f"执行前首页未就绪: {readiness.details}")

            artifacts[f"attempt_{attempt}_before_text"] = save_text_snapshot(
                page, snapshots_dir, case.case_id, attempt, "before"
            )
            artifacts[f"attempt_{attempt}_before_screenshot"] = save_screenshot(
                page, screenshots_dir, case.case_id, attempt, "before"
            )

            page_text = readiness.text or body_text(page)
            assertion_result = assertion(page_text)
            actual = assertion_result if isinstance(assertion_result, str) else assertion_result[0]
            if isinstance(assertion_result, tuple):
                warnings.extend(assertion_result[1])

            artifacts[f"attempt_{attempt}_after_text"] = save_text_snapshot(
                page, snapshots_dir, case.case_id, attempt, "after"
            )
            artifacts[f"attempt_{attempt}_after_screenshot"] = save_screenshot(
                page, screenshots_dir, case.case_id, attempt, "after"
            )
            if attempt > 1:
                warnings.append(f"该用例第 {attempt} 次执行通过，前面共重跑 {attempt - 1} 次。")
            log_step(f"[{case.case_id}] 第 {attempt} 次执行通过。")

            return CaseResult(
                case_id=case.case_id,
                case_name=case.feature,
                status="passed",
                duration_seconds=round(time.monotonic() - start_time, 2),
                expected=case.expected,
                actual=actual,
                page_url=page.url,
                warnings=warnings,
                artifacts=artifacts,
            )
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
            log_step(f"[{case.case_id}] 第 {attempt} 次执行失败: {last_error}")
            warnings.append(f"第 {attempt} 次执行失败: {last_error}")
            artifacts[f"attempt_{attempt}_failure_text"] = save_text_snapshot(
                page, snapshots_dir, case.case_id, attempt, "failure"
            )
            artifacts[f"attempt_{attempt}_failure_screenshot"] = save_screenshot(
                page, screenshots_dir, case.case_id, attempt, "failure"
            )
            artifacts.update(save_failure_logs(report_dir, case.case_id, attempt, console_events, network_errors))

            if attempt > max_retries:
                log_step(f"[{case.case_id}] 已达到最大重跑次数，判定失败。")
                return CaseResult(
                    case_id=case.case_id,
                    case_name=case.feature,
                    status="failed",
                    duration_seconds=round(time.monotonic() - start_time, 2),
                    expected=case.expected,
                    actual=last_error,
                    page_url=page.url,
                    warnings=warnings + summarize_runtime_warnings(console_events, network_errors),
                    artifacts=artifacts,
                )

            # 用例失败后不直接重试断言，而是重新走一遍首页恢复流程，尽量减少偶发白屏和半渲染影响。
            log_step(f"[{case.case_id}] 开始执行失败后的首页恢复。")
            recovery_route, recovery_content = prepare_homepage(
                page, base_url, username, password, route_timeout, content_timeout
            )
            warnings.append(
                f"已完成失败重跑前恢复：首页路由等待 {recovery_route.waited_seconds}s，"
                f"内容等待 {recovery_content.waited_seconds}s。"
            )
            log_step(
                f"[{case.case_id}] 首页恢复完成，路由等待 {recovery_route.waited_seconds}s，"
                f"内容等待 {recovery_content.waited_seconds}s。"
            )

    return CaseResult(
        case_id=case.case_id,
        case_name=case.feature,
        status="failed",
        duration_seconds=round(time.monotonic() - start_time, 2),
        expected=case.expected,
        actual=last_error,
        page_url=page.url,
        warnings=warnings,
        artifacts=artifacts,
    )


def build_failed_case_result(case: CaseDefinition, actual: str, page_url: str) -> CaseResult:
    return CaseResult(
        case_id=case.case_id,
        case_name=case.feature,
        status="failed",
        duration_seconds=0.0,
        expected=case.expected,
        actual=actual,
        page_url=page_url,
        warnings=[],
        artifacts={},
    )


def write_results_json(
    report_dir: Path,
    args: argparse.Namespace,
    cases: list[CaseDefinition],
    results: list[CaseResult],
    route_wait: WaitOutcome | None,
    content_wait: WaitOutcome | None,
) -> Path:
    results_path = report_dir / "results.json"
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "script": "tests/test_homepage_standalone.py",
        "environment": {
            "base_url": args.base_url,
            "excel_path": str(args.excel_path),
            "headless": parse_bool(args.headless),
            "route_timeout_seconds": args.route_timeout,
            "content_timeout_seconds": args.content_timeout,
            "max_retries": args.max_retries,
        },
        "selected_cases": [asdict(case) for case in cases],
        "wait_summary": {
            "route_wait": asdict(route_wait) if route_wait else None,
            "content_wait": asdict(content_wait) if content_wait else None,
        },
        "results": [asdict(result) for result in results],
    }
    write_text_file(results_path, json.dumps(payload, ensure_ascii=False, indent=2))
    return results_path


def write_network_errors(report_dir: Path, network_errors: list[dict[str, object]]) -> Path:
    path = report_dir / "network_errors.json"
    write_text_file(path, json.dumps(network_errors, ensure_ascii=False, indent=2))
    return path


def render_artifacts_html(report_dir: Path, artifacts: dict[str, str]) -> str:
    if not artifacts:
        return "-"

    parts: list[str] = []
    report_dir_resolved = report_dir.resolve()
    cwd_resolved = Path.cwd().resolve()

    for key, value in artifacts.items():
        raw_path = Path(value)
        artifact_path = raw_path if raw_path.is_absolute() else (cwd_resolved / raw_path)
        artifact_path = artifact_path.resolve()

        try:
            relative_path = artifact_path.relative_to(report_dir_resolved)
            src = relative_path.as_posix()
        except ValueError:
            src = artifact_path.as_uri()

        escaped_key = html.escape(key)
        escaped_value = html.escape(value)
        suffix = artifact_path.suffix.lower()
        if suffix == ".png":
            parts.append(
                f"""
                <div class="artifact artifact-image">
                  <div class="artifact-name">{escaped_key}</div>
                  <img
                    src="{src}"
                    alt="{escaped_key}"
                    loading="lazy"
                    class="report-preview-image"
                    data-preview-src="{src}"
                    data-preview-title="{escaped_key}"
                  >
                  <div class="artifact-path">{escaped_value}</div>
                </div>
                """
            )
        else:
            parts.append(
                f"""
                <div class="artifact">
                  <div class="artifact-name">{escaped_key}</div>
                  <div class="artifact-path">{escaped_value}</div>
                </div>
                """
            )

    return "".join(parts)


def write_html_report(
    report_dir: Path,
    args: argparse.Namespace,
    results: list[CaseResult],
    route_wait: WaitOutcome | None,
    content_wait: WaitOutcome | None,
) -> Path:
    report_path = report_dir / "report.html"
    passed_count = sum(1 for result in results if result.status == "passed")
    failed_count = len(results) - passed_count

    rows = []
    for result in results:
        warnings_html = "<br>".join(html.escape(item) for item in result.warnings) or "-"
        artifacts_html = render_artifacts_html(report_dir, result.artifacts)
        rows.append(
            f"""
            <tr class="{result.status}">
              <td>{html.escape(result.case_id)}</td>
              <td>{html.escape(result.case_name)}</td>
              <td>{html.escape(result.status)}</td>
              <td>{result.duration_seconds}</td>
              <td>{html.escape(result.expected)}</td>
              <td>{html.escape(result.actual)}</td>
              <td>{html.escape(result.page_url)}</td>
              <td>{warnings_html}</td>
              <td>{artifacts_html}</td>
            </tr>
            """
        )

    route_summary = html.escape(route_wait.details if route_wait else "-")
    content_summary = html.escape(content_wait.details if content_wait else "-")

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>首页自动化测试报告</title>
  <style>
    body {{ font-family: 'Segoe UI', sans-serif; margin: 24px; background: #f5f7fb; color: #1f2937; }}
    h1 {{ margin-bottom: 8px; }}
    .summary {{ background: #fff; border-radius: 12px; padding: 16px 20px; margin-bottom: 20px; box-shadow: 0 6px 18px rgba(15, 23, 42, 0.08); }}
    table {{ width: 100%; border-collapse: collapse; background: #fff; box-shadow: 0 6px 18px rgba(15, 23, 42, 0.08); }}
    th, td {{ border: 1px solid #dbe2ea; padding: 12px; vertical-align: top; text-align: left; font-size: 14px; }}
    th {{ background: #0f172a; color: #fff; }}
    tr.passed td {{ background: #f0fdf4; }}
    tr.failed td {{ background: #fef2f2; }}
    code {{ background: #eef2ff; padding: 2px 6px; border-radius: 4px; }}
    .artifact {{ margin-bottom: 12px; }}
    .artifact:last-child {{ margin-bottom: 0; }}
    .artifact-name {{ font-weight: 600; margin-bottom: 6px; }}
    .artifact-path {{ color: #475569; font-size: 12px; word-break: break-all; }}
    .artifact-image img {{ display: block; width: 320px; max-width: 100%; border: 1px solid #cbd5e1; border-radius: 8px; margin-bottom: 6px; background: #fff; cursor: zoom-in; }}
    .image-viewer {{ position: fixed; inset: 0; display: none; align-items: center; justify-content: center; background: rgba(15, 23, 42, 0.82); padding: 32px; z-index: 9999; }}
    .image-viewer.is-open {{ display: flex; }}
    .image-viewer-dialog {{ position: relative; max-width: min(96vw, 1600px); max-height: 92vh; }}
    .image-viewer-close {{ position: absolute; top: -18px; right: -18px; width: 40px; height: 40px; border: none; border-radius: 999px; background: #fff; color: #0f172a; font-size: 24px; cursor: pointer; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.25); }}
    .image-viewer img {{ display: block; max-width: 100%; max-height: 82vh; border-radius: 12px; box-shadow: 0 16px 40px rgba(15, 23, 42, 0.38); background: #fff; }}
    .image-viewer-caption {{ margin-top: 10px; color: #e2e8f0; text-align: center; font-size: 14px; }}
  </style>
</head>
<body>
  <h1>首页自动化独立脚本测试报告</h1>
  <div class="summary">
    <p>生成时间: <code>{html.escape(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</code></p>
    <p>目标地址: <code>{html.escape(args.base_url)}</code></p>
    <p>执行结果: 通过 {passed_count} 条，失败 {failed_count} 条</p>
    <p>首页路由等待: {route_summary}</p>
    <p>首页内容等待: {content_summary}</p>
    <p>失败重跑次数上限: {args.max_retries}</p>
  </div>
  <table>
    <thead>
      <tr>
        <th>用例ID</th>
        <th>测试项</th>
        <th>状态</th>
        <th>耗时(s)</th>
        <th>预期</th>
        <th>实际</th>
        <th>页面地址</th>
        <th>警告</th>
        <th>产物</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows)}
    </tbody>
  </table>
  <div class="image-viewer" id="imageViewer" aria-hidden="true">
    <div class="image-viewer-dialog">
      <button type="button" class="image-viewer-close" id="imageViewerClose" aria-label="关闭预览">×</button>
      <img id="imageViewerImg" src="" alt="预览图片">
      <div class="image-viewer-caption" id="imageViewerCaption"></div>
    </div>
  </div>
  <script>
    (function () {{
      const viewer = document.getElementById('imageViewer');
      const viewerImg = document.getElementById('imageViewerImg');
      const viewerCaption = document.getElementById('imageViewerCaption');
      const closeButton = document.getElementById('imageViewerClose');
      const previewImages = document.querySelectorAll('.report-preview-image');

      function closeViewer() {{
        viewer.classList.remove('is-open');
        viewer.setAttribute('aria-hidden', 'true');
        viewerImg.src = '';
        viewerImg.alt = '预览图片';
        viewerCaption.textContent = '';
      }}

      previewImages.forEach(function (img) {{
        img.addEventListener('click', function () {{
          const src = img.getAttribute('data-preview-src') || img.getAttribute('src') || '';
          const title = img.getAttribute('data-preview-title') || img.getAttribute('alt') || '';
          viewerImg.src = src;
          viewerImg.alt = title;
          viewerCaption.textContent = title;
          viewer.classList.add('is-open');
          viewer.setAttribute('aria-hidden', 'false');
        }});
      }});

      closeButton.addEventListener('click', closeViewer);
      viewer.addEventListener('click', function (event) {{
        if (event.target === viewer) {{
          closeViewer();
        }}
      }});
      document.addEventListener('keydown', function (event) {{
        if (event.key === 'Escape' && viewer.classList.contains('is-open')) {{
          closeViewer();
        }}
      }});
    }})();
  </script>
</body>
</html>
"""
    write_text_file(report_path, html_content)
    return report_path


def run() -> int:
    args = parse_args()
    excel_path = Path(args.excel_path)
    report_dir = build_report_dir(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    log_step("开始执行首页自动化脚本。")
    log_step(
        f"运行参数: headless={parse_bool(args.headless)}, route_timeout={args.route_timeout}s, "
        f"content_timeout={args.content_timeout}s, max_retries={args.max_retries}。"
    )

    console_log_path = report_dir / "console.log"
    write_text_file(console_log_path, "")

    cases = load_home_cases(excel_path)
    log_step(f"已加载首页测试用例 {len(cases)} 条: {', '.join(case.case_id for case in cases)}。")
    console_events: list[dict[str, object]] = []
    network_errors: list[dict[str, object]] = []
    results: list[CaseResult] = []
    route_wait: WaitOutcome | None = None
    content_wait: WaitOutcome | None = None

    with sync_playwright() as playwright:
        log_step("正在启动 Playwright 浏览器。")
        browser = playwright.chromium.launch(channel="chrome", headless=parse_bool(args.headless))
        context = browser.new_context(ignore_https_errors=True, viewport={"width": 1600, "height": 900})
        page = context.new_page()
        log_step("浏览器已启动，开始绑定日志监听。")

        def on_console(message) -> None:  # type: ignore[no-untyped-def]
            # 控制台信息写入内存和文件两份，既便于最终报告汇总，也便于单独排错。
            entry = {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "level": message.type,
                "text": message.text,
            }
            console_events.append(entry)
            append_console_log(console_log_path, message.type, message.text)

        def on_pageerror(error) -> None:  # type: ignore[no-untyped-def]
            entry = {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "level": "pageerror",
                "text": str(error),
            }
            console_events.append(entry)
            append_console_log(console_log_path, "pageerror", str(error))

        def on_response(response: Response) -> None:
            if response.status < 400:
                return
            # 只抓失败响应，避免网络日志过大；报告里按 warning 呈现最近失败请求。
            network_errors.append(
                {
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "status": response.status,
                    "url": response.url,
                }
            )

        page.on("console", on_console)
        page.on("pageerror", on_pageerror)
        page.on("response", on_response)

        try:
            last_prepare_error = ""
            for attempt in range(1, args.max_retries + 2):
                try:
                    log_step(f"开始第 {attempt} 次首页准备。")
                    # 整体执行开始前也允许首页准备失败重试，适配登录后长时间空白的场景。
                    route_wait, content_wait = prepare_homepage(
                        page,
                        args.base_url,
                        args.username,
                        args.password,
                        args.route_timeout,
                        args.content_timeout,
                    )
                    if attempt > 1:
                        append_console_log(console_log_path, "info", f"首页准备在第 {attempt} 次尝试后成功。")
                    log_step(
                        f"首页准备成功，路由等待 {route_wait.waited_seconds}s，"
                        f"内容等待 {content_wait.waited_seconds}s。"
                    )
                    break
                except Exception as exc:  # noqa: BLE001
                    last_prepare_error = str(exc)
                    log_step(f"第 {attempt} 次首页准备失败: {last_prepare_error}")
                    failure_path = report_dir / "screenshots" / f"prepare_attempt_{attempt}_failure.png"
                    capture_screenshot(page, failure_path, f"prepare attempt {attempt} failure")
                    if attempt > args.max_retries:
                        raise RuntimeError(last_prepare_error) from exc
            else:
                raise RuntimeError(last_prepare_error or "首页准备失败")

            case_map = {case.case_id: case for case in cases}
            log_step("开始执行测试用例 TC-HOME-007。")
            results.append(
                run_case_with_retry(
                    page,
                    case_map["TC-HOME-007"],
                    report_dir,
                    console_events,
                    network_errors,
                    assert_home_generation,
                    args.base_url,
                    args.username,
                    args.password,
                    args.route_timeout,
                    args.content_timeout,
                    args.max_retries,
                )
            )
            log_step("开始执行测试用例 TC-HOME-011。")
            results.append(
                run_case_with_retry(
                    page,
                    case_map["TC-HOME-011"],
                    report_dir,
                    console_events,
                    network_errors,
                    assert_turbine_and_pv_sections,
                    args.base_url,
                    args.username,
                    args.password,
                    args.route_timeout,
                    args.content_timeout,
                    args.max_retries,
                )
            )
            log_step("开始执行测试用例 TC-HOME-014。")
            results.append(
                run_case_with_retry(
                    page,
                    case_map["TC-HOME-014"],
                    report_dir,
                    console_events,
                    network_errors,
                    lambda text: assert_energy_reduction(text, console_events, network_errors),
                    args.base_url,
                    args.username,
                    args.password,
                    args.route_timeout,
                    args.content_timeout,
                    args.max_retries,
                )
            )
        except Exception as exc:  # noqa: BLE001
            log_step(f"执行过程中发生致命错误: {exc}")
            failure_screenshot = report_dir / "screenshots" / "run_level_failure.png"
            capture_screenshot(page, failure_screenshot, "run level failure")

            failure_message = str(exc)
            if not results:
                results = [build_failed_case_result(case, failure_message, page.url) for case in cases]
            else:
                for case in cases[len(results):]:
                    results.append(build_failed_case_result(case, failure_message, page.url))
        finally:
            log_step("正在关闭浏览器上下文。")
            context.close()
            browser.close()

    log_step("开始写入测试报告与附属产物。")
    write_network_errors(report_dir, network_errors)
    write_results_json(report_dir, args, cases, results, route_wait, content_wait)
    write_html_report(report_dir, args, results, route_wait, content_wait)

    passed_count = sum(1 for result in results if result.status == "passed")
    failed_count = len(results) - passed_count
    log_step(f"测试完成: 通过 {passed_count} 条, 失败 {failed_count} 条。")
    log_step(f"报告目录: {report_dir}")
    return 0 if failed_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(run())
