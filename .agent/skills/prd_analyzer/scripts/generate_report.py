import pandas as pd
import json
import sys
import os
import numpy as np
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel(data, output_file=None):
    """
    将 PRD 分析 JSON 转换为标准的 3-Sheet Excel 报告。
    支持多种 JSON 结构变体，增强鲁棒性。
    """
    project_info = data.get("project_info", {})
    if isinstance(project_info, str): project_info = {"name": project_info}
    project_name = project_info.get("name") or project_info.get("projectName") or "未命名项目"
    
    if output_file is None:
        output_file = os.path.join("output", f"{project_name}-测试报告.xlsx")

    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    # 1. 动态确定测试用例 Sheet 名称
    preferred_name = data.get("preferred_sheet_name")
    modules_data = data.get("modules", [])
    
    if isinstance(modules_data, dict):
        # 兼容字典结构 {"module1": {...}, "module2": {...}}
        processed_modules = []
        for m_key, m_val in modules_data.items():
            if isinstance(m_val, dict):
                if "module_name" not in m_val: m_val["module_name"] = m_key
                processed_modules.append(m_val)
        modules_data = processed_modules

    if not preferred_name and len(modules_data) == 1:
        preferred_name = modules_data[0].get("module_name")
    
    case_sheet_name = preferred_name or "功能测试用例"

    # 1. 数据映射配置 (支持中英文 Key)
    mapping = {
        case_sheet_name: {
            "cols": ["所属模块", "测试用例 ID", "测试项", "优先级", "重要程度", "测试步骤", "预期结果", "测试结果", "备注"],
            "map": {
                "测试用例 ID": ["id", "caseId", "用例ID", "测试用例ID"],
                "测试项": ["title", "scenario", "testItem", "测试场景"],
                "优先级": ["priority"],
                "重要程度": ["importance", "level", "P等级", "重要性"],
                "测试步骤": ["steps", "procedure", "testSteps"],
                "预期结果": ["expected", "expectedResult", "expectedResults"],
                "测试结果": ["actual", "testResult", "result"],
                "备注": ["remark", "notes", "description"]
            }
        },
        "需求测试点": {
            "cols": ["所属模块", "id", "description", "category"],
            "map": {
                "id": ["ID", "testPointId", "测试点ID"],
                "description": ["desc", "detail", "测试点描述", "描述"],
                "category": ["type", "分类", "测试分类"]
            }
        },
        "测试数据准备": {
            "cols": ["所属模块", "字段", "数据示例", "使用场景"],
            "map": {
                "字段": ["field", "fieldName", "参数"],
                "数据示例": ["example", "data", "value", "数据值"],
                "使用场景": ["scenario", "usage", "context", "说明"]
            }
        }
    }

    # 2. 聚合数据
    aggregated = {k: [] for k in mapping}
    
    for module in modules_data:
        if not isinstance(module, dict): continue
        m_name = module.get("module_name") or module.get("name") or "默认模块"
        
        # 处理每个 Sheet
        data_keys = {
            case_sheet_name: ["cases", "testCases", "test_cases"],
            "需求测试点": ["test_points", "testPoints", "points"],
            "测试数据准备": ["test_data", "testData", "data"]
        }
        
        for sheet_name, keys in data_keys.items():
            items = []
            for k in keys:
                if k in module:
                    items = module[k]
                    break
            
            if not isinstance(items, list): continue
            
            sheet_map = mapping[sheet_name]["map"]
            for item in items:
                if not isinstance(item, dict): continue
                row = {"所属模块": m_name}
                # 按照映射填充
                for target_col, synonyms in sheet_map.items():
                    val = item.get(target_col)
                    if val is None:
                        for s in synonyms:
                            if s in item:
                                val = item[s]
                                break
                    row[target_col] = val if val is not None else ""
                aggregated[sheet_name].append(row)

    # 3. 写入 Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, sheet_config in mapping.items():
            cols = sheet_config["cols"]
            df = pd.DataFrame(aggregated[sheet_name])
            
            # 补全列并排序
            for c in cols:
                if c not in df.columns:
                    df[c] = ""
            df = df[cols].fillna("")

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 样式与排版
            ws = writer.sheets[sheet_name]
            
            # 表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 边框与对齐
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(cols)):
                for cell in row:
                    cell.border = border
                    if cell.row > 1:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # 自动调整列宽 (根据配置调整)
            width_map = {
                "所属模块": 15, "测试用例 ID": 15, "优先级": 10, "重要程度": 10,
                "测试步骤": 45, "预期结果": 45, "测试点描述": 40, "description": 40
            }
            for i, col in enumerate(cols):
                w = width_map.get(col, 20)
                ws.column_dimensions[get_column_letter(i+1)].width = w

            # 冻结首行与筛选
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    print(f"✅ GPT 测试报告已成功生成: {os.path.abspath(output_file)}")

if __name__ == "__main__":
    try:
        input_path = sys.argv[1] if len(sys.argv) > 1 else None
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        
        if input_path:
            with open(input_path, 'r', encoding='utf-8') as f:
                raw_data = json.load(f)
        else:
            raw_data = json.load(sys.stdin)
            
        generate_excel(raw_data, output_path)
    except Exception as e:
        print(f"❌ Excel 生成异常: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
